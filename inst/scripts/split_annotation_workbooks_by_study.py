from __future__ import annotations

import json
from pathlib import Path
import sys
from typing import Iterable
import xml.etree.ElementTree as ET
from zipfile import ZipFile
import csv

from openpyxl import load_workbook
import pandas as pd


ROOT = Path("/Users/nalinarora/Desktop/Microbiome_Metabolome_Curated_Studies")
STUDY_INDEX_PATH = ROOT / "inst" / "extdata" / "study_index.csv"
MAPPED_WORKBOOK_PATH = ROOT / "all_studies_mapped_stage4.xlsx"
DRUG_WORKBOOK_PATH = ROOT / "Humannet_Library_V1_drug_annotations.xlsx"
OUTPUT_DIR = ROOT / "study_annotation_outputs"

MAPPED_COLUMNS = [
    "HuMANet_ID",
    "Query_Name",
    "Database_Source",
    "InChIKey",
    "HMDB_ID",
    "PubChem_CID",
    "KEGG_ID",
    "ChEBI_ID",
    "Molecular_Formula",
    "Exact_Mass",
    "Super_Class",
    "Main_Class",
    "Sub_Class",
    "Standardized_Name",
    "SMILES",
    "Synonyms",
]

MULTI_VALUE_SHEETS = ["species", "disease", "pathways"]
DRUG_SHEETS = ["drugbank_hits", "drugcentral_hits"]
REMOVED_STUDY_PREFIXES = ("LifelinesDEEP",)
REMOVED_STUDIES = {"Diener_2022"}


def split_pipe(value: object) -> list[str]:
    if pd.isna(value):
        return []
    return [part.strip() for part in str(value).split("|") if part.strip()]


def active_studies() -> pd.DataFrame:
    idx = pd.read_csv(STUDY_INDEX_PATH)
    idx = idx[~idx["study"].isin(REMOVED_STUDIES)].copy()
    idx = idx[~idx["study"].str.startswith(REMOVED_STUDY_PREFIXES)].copy()
    idx["study_folder"] = idx["study"] + "_CuratedMetabolome"
    return idx


def filter_known_studies(parts: Iterable[str], study_folders: set[str]) -> list[str]:
    return [part for part in parts if part in study_folders]


def study_output_dir(study: str) -> Path:
    path = OUTPUT_DIR / study
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)


def workbook_headers(workbook_path: Path, sheet: str) -> list[str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        return ["" if value is None else str(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    finally:
        wb.close()


def read_csv_hman_ids(path: Path, split_values: bool) -> set[str]:
    if not path.exists():
        return set()

    seen_ids: set[str] = set()
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            value = row.get("HuMANet_ID")
            if not value:
                continue
            if split_values:
                seen_ids.update(split_pipe(value))
            else:
                seen_ids.add(value)
    return seen_ids


def excel_column_label(cell_ref: str) -> str:
    letters: list[str] = []
    for char in cell_ref:
        if char.isalpha():
            letters.append(char)
        else:
            break
    return "".join(letters)


def load_shared_strings(xlsx_path: Path) -> list[str]:
    with ZipFile(xlsx_path) as archive:
        try:
            with archive.open("xl/sharedStrings.xml") as handle:
                strings: list[str] = []
                for _, elem in ET.iterparse(handle, events=("end",)):
                    if elem.tag.endswith("}si"):
                        parts = [text for text in elem.itertext()]
                        strings.append("".join(parts))
                        elem.clear()
                return strings
        except KeyError:
            return []


def parse_sheet_xml_rows(
    xlsx_path: Path,
    sheet_xml_path: str,
    shared_strings: list[str] | None = None,
) -> Iterable[dict[str, object]]:
    shared_strings = shared_strings or []

    with ZipFile(xlsx_path) as archive:
        with archive.open(sheet_xml_path) as handle:
            header_by_col: dict[str, str] = {}

            row_index = 0
            for _, elem in ET.iterparse(handle, events=("end",)):
                if not elem.tag.endswith("}row"):
                    continue

                row_values: dict[str, object] = {}

                for cell in elem:
                    if not cell.tag.endswith("}c"):
                        continue

                    cell_ref = cell.attrib.get("r", "")
                    col_label = excel_column_label(cell_ref)
                    cell_type = cell.attrib.get("t")
                    value = None

                    if cell_type == "inlineStr":
                        texts = [text for text in cell.itertext()]
                        value = "".join(texts) if texts else None
                    else:
                        value_elem = next((child for child in cell if child.tag.endswith("}v")), None)
                        if value_elem is not None:
                            raw = value_elem.text
                            if cell_type == "s" and raw is not None:
                                value = shared_strings[int(raw)]
                            else:
                                value = raw

                    if not header_by_col:
                        row_values[col_label] = value
                    elif col_label in header_by_col:
                        row_values[header_by_col[col_label]] = value

                if not header_by_col:
                    header_by_col = {
                        col_label: str(value)
                        for col_label, value in row_values.items()
                        if value is not None
                    }
                elif row_values:
                    row_index += 1
                    if row_index % 5000 == 0:
                        print(f"{sheet_xml_path}: processed {row_index} rows", flush=True)
                    yield row_values

                elem.clear()


def sheet_xml_headers(
    xlsx_path: Path,
    sheet_xml_path: str,
    shared_strings: list[str] | None = None,
) -> list[str]:
    shared_strings = shared_strings or []

    with ZipFile(xlsx_path) as archive:
        with archive.open(sheet_xml_path) as handle:
            for _, elem in ET.iterparse(handle, events=("end",)):
                if not elem.tag.endswith("}row"):
                    continue

                row_values: dict[str, str] = {}
                for cell in elem:
                    if not cell.tag.endswith("}c"):
                        continue

                    cell_ref = cell.attrib.get("r", "")
                    col_label = excel_column_label(cell_ref)
                    cell_type = cell.attrib.get("t")
                    value = None

                    if cell_type == "inlineStr":
                        texts = [text for text in cell.itertext()]
                        value = "".join(texts) if texts else None
                    else:
                        value_elem = next((child for child in cell if child.tag.endswith("}v")), None)
                        if value_elem is not None:
                            raw = value_elem.text
                            if cell_type == "s" and raw is not None:
                                value = shared_strings[int(raw)]
                            else:
                                value = raw

                    if value is not None:
                        row_values[col_label] = str(value)

                elem.clear()
                return [row_values[key] for key in sorted(row_values)]

    return []


def create_mapped_outputs(
    mapped_df: pd.DataFrame,
    studies_df: pd.DataFrame,
) -> tuple[dict[str, pd.DataFrame], list[dict[str, object]]]:
    mapped_outputs: dict[str, pd.DataFrame] = {}
    summary: list[dict[str, object]] = []
    active_folders = set(studies_df["study_folder"])

    filtered = mapped_df[mapped_df["Study_Folder"].isin(active_folders)].copy()

    for row in studies_df.itertuples(index=False):
        study = row.study
        study_folder = row.study_folder
        study_df = filtered.loc[filtered["Study_Folder"] == study_folder, MAPPED_COLUMNS].copy()
        mapped_outputs[study] = study_df
        write_csv(study_df, study_output_dir(study) / "mapped.csv")
        summary.append(
            {
                "study": study,
                "sheet": "mapped",
                "rows": int(len(study_df)),
                "unique_hman_ids": int(study_df["HuMANet_ID"].nunique()),
            }
        )

    return mapped_outputs, summary


def trim_multi_value_row(row: pd.Series, study_folder: str, valid_hman_ids: set[str]) -> dict[str, object] | None:
    hman_ids = split_pipe(row["HuMANet_ID"])
    query_names = split_pipe(row["Query_Name"])
    study_folders = split_pipe(row["Study_Folder"])

    if not study_folders or study_folder not in study_folders:
        return None

    kept_ids: list[str] = []
    kept_names: list[str] = []

    for idx, hman_id in enumerate(hman_ids):
        if hman_id in valid_hman_ids:
            kept_ids.append(hman_id)
            if idx < len(query_names):
                kept_names.append(query_names[idx])

    if not kept_ids:
        return None

    trimmed = row.drop(labels=["Study_Folder"]).to_dict()
    trimmed["HuMANet_ID"] = "|".join(kept_ids)
    if query_names:
        trimmed["Query_Name"] = "|".join(kept_names) if kept_names else row["Query_Name"]
    return trimmed


def create_multi_value_outputs(
    workbook: pd.ExcelFile,
    studies_df: pd.DataFrame,
    mapped_outputs: dict[str, pd.DataFrame],
    sheets: list[str] | None = None,
) -> tuple[dict[str, dict[str, pd.DataFrame]], list[dict[str, object]]]:
    target_sheets = sheets or MULTI_VALUE_SHEETS
    outputs: dict[str, dict[str, pd.DataFrame]] = {sheet: {} for sheet in target_sheets}
    summary: list[dict[str, object]] = []
    folder_to_study = dict(zip(studies_df["study_folder"], studies_df["study"]))
    mapped_id_sets = {
        study: set(df["HuMANet_ID"])
        for study, df in mapped_outputs.items()
    }

    for sheet in target_sheets:
        df = workbook.parse(sheet)
        rows_by_study: dict[str, list[dict[str, object]]] = {
            study: [] for study in studies_df["study"]
        }

        for _, record in df.iterrows():
            study_folders = filter_known_studies(
                split_pipe(record["Study_Folder"]),
                set(folder_to_study),
            )
            if not study_folders:
                continue

            hman_ids = split_pipe(record["HuMANet_ID"])
            query_names = split_pipe(record["Query_Name"])

            for study_folder in study_folders:
                study = folder_to_study[study_folder]
                valid_hman_ids = mapped_id_sets[study]
                kept_ids: list[str] = []
                kept_names: list[str] = []

                for idx, hman_id in enumerate(hman_ids):
                    if hman_id in valid_hman_ids:
                        kept_ids.append(hman_id)
                        if idx < len(query_names):
                            kept_names.append(query_names[idx])

                if not kept_ids:
                    continue

                trimmed = record.drop(labels=["Study_Folder"]).to_dict()
                trimmed["HuMANet_ID"] = "|".join(kept_ids)
                if query_names:
                    trimmed["Query_Name"] = "|".join(kept_names) if kept_names else record["Query_Name"]
                rows_by_study[study].append(trimmed)

        for row in studies_df.itertuples(index=False):
            study = row.study
            out_df = pd.DataFrame(rows_by_study[study])
            outputs[sheet][study] = out_df
            write_csv(out_df, study_output_dir(study) / f"{sheet}.csv")

            unique_ids = set()
            if not out_df.empty:
                for value in out_df["HuMANet_ID"]:
                    unique_ids.update(split_pipe(value))

            summary.append(
                {
                    "study": study,
                    "sheet": sheet,
                    "rows": int(len(out_df)),
                    "unique_hman_ids": int(len(unique_ids)),
                }
            )

    return outputs, summary


def create_multi_value_outputs_streaming(
    workbook_path: Path,
    studies_df: pd.DataFrame,
    mapped_outputs: dict[str, pd.DataFrame],
    sheets: list[str] | None = None,
) -> tuple[dict[str, dict[str, pd.DataFrame]], list[dict[str, object]]]:
    target_sheets = sheets or MULTI_VALUE_SHEETS
    outputs: dict[str, dict[str, pd.DataFrame]] = {}
    summary: list[dict[str, object]] = []
    folder_to_study = dict(zip(studies_df["study_folder"], studies_df["study"]))
    active_folders = set(folder_to_study)
    mapped_id_sets = {
        study: set(df["HuMANet_ID"])
        for study, df in mapped_outputs.items()
    }
    shared_strings = load_shared_strings(workbook_path)
    sheet_xml_map = {
        "species": "xl/worksheets/sheet2.xml",
        "disease": "xl/worksheets/sheet3.xml",
        "pathways": "xl/worksheets/sheet4.xml",
    }

    for sheet in target_sheets:
        xml_path = sheet_xml_map[sheet]
        header = [col for col in sheet_xml_headers(workbook_path, xml_path, shared_strings) if col != "Study_Folder"]
        row_counts = {study: 0 for study in studies_df["study"]}
        unique_ids_by_study = {study: set() for study in studies_df["study"]}
        handles: dict[str, object] = {}
        writers: dict[str, csv.DictWriter] = {}

        for study in studies_df["study"]:
            handle = (study_output_dir(study) / f"{sheet}.csv").open("w", newline="", encoding="utf-8")
            writer = csv.DictWriter(handle, fieldnames=header)
            writer.writeheader()
            handles[study] = handle
            writers[study] = writer

        try:
            for record in parse_sheet_xml_rows(workbook_path, xml_path, shared_strings):
                study_folders = filter_known_studies(
                    split_pipe(record.get("Study_Folder")),
                    active_folders,
                )
                if not study_folders:
                    continue

                hman_ids = split_pipe(record.get("HuMANet_ID"))
                query_names = split_pipe(record.get("Query_Name"))

                for study_folder in study_folders:
                    study = folder_to_study[study_folder]
                    valid_hman_ids = mapped_id_sets[study]
                    kept_ids: list[str] = []
                    kept_names: list[str] = []

                    for idx, hman_id in enumerate(hman_ids):
                        if hman_id in valid_hman_ids:
                            kept_ids.append(hman_id)
                            if idx < len(query_names):
                                kept_names.append(query_names[idx])

                    if not kept_ids:
                        continue

                    trimmed = {k: record.get(k) for k in header}
                    trimmed["HuMANet_ID"] = "|".join(kept_ids)
                    if "Query_Name" in trimmed and query_names:
                        trimmed["Query_Name"] = "|".join(kept_names) if kept_names else record.get("Query_Name")

                    writers[study].writerow(trimmed)
                    row_counts[study] += 1
                    unique_ids_by_study[study].update(kept_ids)
        finally:
            for handle in handles.values():
                handle.close()

        for study in studies_df["study"]:
            summary.append(
                {
                    "study": study,
                    "sheet": sheet,
                    "rows": int(row_counts[study]),
                    "unique_hman_ids": int(len(unique_ids_by_study[study])),
                }
            )

    return outputs, summary


def create_drug_outputs(
    workbook: pd.ExcelFile,
    studies_df: pd.DataFrame,
    mapped_outputs: dict[str, pd.DataFrame],
    sheets: list[str] | None = None,
) -> tuple[dict[str, dict[str, pd.DataFrame]], list[dict[str, object]]]:
    target_sheets = sheets or DRUG_SHEETS
    outputs: dict[str, dict[str, pd.DataFrame]] = {sheet: {} for sheet in target_sheets}
    summary: list[dict[str, object]] = []
    folder_to_study = dict(zip(studies_df["study_folder"], studies_df["study"]))
    mapped_id_sets = {
        study: set(df["HuMANet_ID"])
        for study, df in mapped_outputs.items()
    }

    for sheet in target_sheets:
        df = workbook.parse(sheet)
        rows_by_study: dict[str, list[dict[str, object]]] = {
            study: [] for study in studies_df["study"]
        }

        for _, record in df.iterrows():
            study_folder = record["Study_Folder"]
            if study_folder not in folder_to_study:
                continue
            study = folder_to_study[study_folder]
            if record["HuMANet_ID"] not in mapped_id_sets[study]:
                continue
            rows_by_study[study].append(record.drop(labels=["Study_Folder"]).to_dict())

        for row in studies_df.itertuples(index=False):
            study = row.study
            out_df = pd.DataFrame(rows_by_study[study])
            outputs[sheet][study] = out_df
            write_csv(out_df, study_output_dir(study) / f"{sheet}.csv")
            summary.append(
                {
                    "study": study,
                    "sheet": sheet,
                    "rows": int(len(out_df)),
                    "unique_hman_ids": int(out_df["HuMANet_ID"].nunique()),
                }
            )

    return outputs, summary


def create_drug_outputs_streaming(
    workbook_path: Path,
    studies_df: pd.DataFrame,
    mapped_outputs: dict[str, pd.DataFrame],
    sheets: list[str] | None = None,
) -> tuple[dict[str, dict[str, pd.DataFrame]], list[dict[str, object]]]:
    target_sheets = sheets or DRUG_SHEETS
    outputs: dict[str, dict[str, pd.DataFrame]] = {}
    summary: list[dict[str, object]] = []
    folder_to_study = dict(zip(studies_df["study_folder"], studies_df["study"]))
    mapped_id_sets = {
        study: set(df["HuMANet_ID"])
        for study, df in mapped_outputs.items()
    }
    sheet_xml_map = {
        "drugbank_hits": "xl/worksheets/sheet1.xml",
        "drugcentral_hits": "xl/worksheets/sheet2.xml",
    }

    for sheet in target_sheets:
        xml_path = sheet_xml_map[sheet]
        header = [col for col in sheet_xml_headers(workbook_path, xml_path) if col != "Study_Folder"]
        row_counts = {study: 0 for study in studies_df["study"]}
        unique_ids_by_study = {study: set() for study in studies_df["study"]}
        handles: dict[str, object] = {}
        writers: dict[str, csv.DictWriter] = {}

        for study in studies_df["study"]:
            handle = (study_output_dir(study) / f"{sheet}.csv").open("w", newline="", encoding="utf-8")
            writer = csv.DictWriter(handle, fieldnames=header)
            writer.writeheader()
            handles[study] = handle
            writers[study] = writer

        try:
            for record in parse_sheet_xml_rows(workbook_path, xml_path):
                study_folder = record.get("Study_Folder")
                if study_folder not in folder_to_study:
                    continue
                study = folder_to_study[study_folder]
                hman_id = record.get("HuMANet_ID")
                if hman_id not in mapped_id_sets[study]:
                    continue

                writers[study].writerow({k: record.get(k) for k in header})
                row_counts[study] += 1
                unique_ids_by_study[study].add(hman_id)
        finally:
            for handle in handles.values():
                handle.close()

        for study in studies_df["study"]:
            summary.append(
                {
                    "study": study,
                    "sheet": sheet,
                    "rows": int(row_counts[study]),
                    "unique_hman_ids": int(len(unique_ids_by_study[study])),
                }
            )

    return outputs, summary


def check_consistency(
    studies_df: pd.DataFrame,
    mapped_outputs: dict[str, pd.DataFrame],
    multi_outputs: dict[str, dict[str, pd.DataFrame]] | None = None,
    drug_outputs: dict[str, dict[str, pd.DataFrame]] | None = None,
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    multi_outputs = multi_outputs or {}
    drug_outputs = drug_outputs or {}

    for row in studies_df.itertuples(index=False):
        study = row.study
        mapped_ids = set(mapped_outputs[study]["HuMANet_ID"])

        for sheet in MULTI_VALUE_SHEETS:
            out_df = multi_outputs.get(sheet, {}).get(study)
            if out_df is None:
                path = study_output_dir(study) / f"{sheet}.csv"
                seen_ids = read_csv_hman_ids(path, split_values=True)
            else:
                seen_ids = set()
                if not out_df.empty:
                    for value in out_df["HuMANet_ID"]:
                        seen_ids.update(split_pipe(value))

            unexpected = sorted(seen_ids - mapped_ids)
            missing = sorted(mapped_ids - seen_ids)
            records.append(
                {
                    "study": study,
                    "sheet": sheet,
                    "mapped_hman_ids": len(mapped_ids),
                    "output_hman_ids": len(seen_ids),
                    "unexpected_hman_ids": len(unexpected),
                    "missing_hman_ids": len(missing),
                    "status": "ok" if not unexpected else "unexpected_ids",
                }
            )

        for sheet in DRUG_SHEETS:
            out_df = drug_outputs.get(sheet, {}).get(study)
            if out_df is None:
                path = study_output_dir(study) / f"{sheet}.csv"
                seen_ids = read_csv_hman_ids(path, split_values=False)
            else:
                seen_ids = set(out_df["HuMANet_ID"]) if not out_df.empty else set()
            unexpected = sorted(seen_ids - mapped_ids)
            records.append(
                {
                    "study": study,
                    "sheet": sheet,
                    "mapped_hman_ids": len(mapped_ids),
                    "output_hman_ids": len(seen_ids),
                    "unexpected_hman_ids": len(unexpected),
                    "missing_hman_ids": pd.NA,
                    "status": "ok" if not unexpected else "unexpected_ids",
                }
            )

    return pd.DataFrame(records)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    mode = sys.argv[1] if len(sys.argv) > 1 else "all"

    studies_df = active_studies()
    mapped_workbook = pd.ExcelFile(MAPPED_WORKBOOK_PATH)
    drug_workbook = pd.ExcelFile(DRUG_WORKBOOK_PATH)

    if mode == "resume_remaining":
        mapped_outputs = {
            study: pd.read_csv(study_output_dir(study) / "mapped.csv")
            for study in studies_df["study"]
        }
        mapped_summary = []
        multi_outputs, multi_summary = create_multi_value_outputs_streaming(
            MAPPED_WORKBOOK_PATH,
            studies_df,
            mapped_outputs,
            sheets=["pathways"],
        )
        drug_outputs, drug_summary = create_drug_outputs_streaming(
            DRUG_WORKBOOK_PATH,
            studies_df,
            mapped_outputs,
            sheets=DRUG_SHEETS,
        )
    elif mode == "drugs_and_reports":
        mapped_outputs = {
            study: pd.read_csv(study_output_dir(study) / "mapped.csv")
            for study in studies_df["study"]
        }
        mapped_summary = []
        multi_outputs = {}
        multi_summary = []
        drug_outputs, drug_summary = create_drug_outputs_streaming(
            DRUG_WORKBOOK_PATH,
            studies_df,
            mapped_outputs,
            sheets=DRUG_SHEETS,
        )
    elif mode == "pathways_and_reports":
        mapped_outputs = {
            study: pd.read_csv(study_output_dir(study) / "mapped.csv")
            for study in studies_df["study"]
        }
        mapped_summary = []
        multi_outputs, multi_summary = create_multi_value_outputs_streaming(
            MAPPED_WORKBOOK_PATH,
            studies_df,
            mapped_outputs,
            sheets=["pathways"],
        )
        drug_outputs = {}
        drug_summary = []
    else:
        mapped_outputs, mapped_summary = create_mapped_outputs(
            mapped_workbook.parse("mapped"),
            studies_df,
        )
        multi_outputs, multi_summary = create_multi_value_outputs(
            mapped_workbook,
            studies_df,
            mapped_outputs,
        )
        drug_outputs, drug_summary = create_drug_outputs(
            drug_workbook,
            studies_df,
            mapped_outputs,
        )

    summary_df = pd.DataFrame(mapped_summary + multi_summary + drug_summary)
    if mode in {"resume_remaining", "drugs_and_reports", "pathways_and_reports"}:
        existing_summary_path = OUTPUT_DIR / "annotation_split_summary.csv"
        if existing_summary_path.exists():
            existing_summary = pd.read_csv(existing_summary_path)
            summary_df = pd.concat([existing_summary, summary_df], ignore_index=True)
            summary_df = summary_df.drop_duplicates(subset=["study", "sheet"], keep="last")

    consistency_df = check_consistency(
        studies_df,
        mapped_outputs,
        multi_outputs,
        drug_outputs,
    )

    write_csv(summary_df, OUTPUT_DIR / "annotation_split_summary.csv")
    write_csv(consistency_df, OUTPUT_DIR / "annotation_hman_consistency.csv")

    report = {
        "n_studies": int(len(studies_df)),
        "removed_studies": sorted(list(REMOVED_STUDIES)),
        "removed_prefixes": list(REMOVED_STUDY_PREFIXES),
        "consistency_status_counts": consistency_df["status"].value_counts(dropna=False).to_dict(),
        "summary_rows": int(len(summary_df)),
    }
    (OUTPUT_DIR / "annotation_split_report.json").write_text(
        json.dumps(report, indent=2),
        encoding="utf-8",
    )

    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
